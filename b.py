import requests, json
import datetime,time
from datetime import datetime

import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

headers = {
    "User-Agent":
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36 Edge/18.19582"
}

bk = openpyxl.load_workbook('C:/Users/masrafe/Desktop/selenium/Excel.xlsx')
s = bk.active
b = s.cell (row = 3, column = 3).value
a = 'http://google.com/complete/search?client=chrome&q='+ b
response = requests.get(a, headers=headers)
elements = []
for result in json.loads(response.text)[1]:
    elements.append(result)

dt = datetime.now()
day =dt.strftime('%A')
date_time = datetime.now()

long = max(elements, key=len) 
short = min(elements,key=len)


workbook = openpyxl.load_workbook('C:/Users/masrafe/Desktop/selenium/Excel.xlsx')
sheet = workbook.active
current_row = sheet.max_row
sheet.cell(4,5).value = short
sheet.cell(4,4).value = long
sheet.cell(4,6).value = day
sheet.cell(4,7).value = date_time


os.chmod('C:/Users/masrafe/Desktop/selenium/Excel.xlsx', 0o777)

workbook.save('C:/Users/masrafe/Desktop/selenium/Excel.xlsx')

